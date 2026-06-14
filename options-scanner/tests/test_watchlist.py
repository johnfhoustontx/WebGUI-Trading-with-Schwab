"""Tests for the scan watchlist module."""

import os

import openpyxl

import watchlist


def _make_xlsx(path, col_a):
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, v in enumerate(col_a, start=1):
        ws.cell(row=i, column=1, value=v)
        ws.cell(row=i, column=2, value=f"Name ({v})")
    wb.save(path)


def test_union_base_and_column_a_order(tmp_path):
    p = tmp_path / "Top 20.xlsx"
    _make_xlsx(p, ["NVDA", "TSLA", "AAPL"])
    syms = watchlist.read_symbols(p)
    assert syms == ["$SPX", "SPY", "QQQ", "NVDA", "TSLA", "AAPL"]


def test_dedupe_when_indices_in_sheet(tmp_path):
    p = tmp_path / "Top 20.xlsx"
    _make_xlsx(p, ["$SPX", "SPY", "QQQ", "NVDA"])
    syms = watchlist.read_symbols(p)
    assert syms == ["$SPX", "SPY", "QQQ", "NVDA"]


def test_normalizes_whitespace_and_case(tmp_path):
    p = tmp_path / "Top 20.xlsx"
    _make_xlsx(p, [" nvda ", "tsla", "", None, "AAPL"])
    syms = watchlist.read_symbols(p)
    assert syms == ["$SPX", "SPY", "QQQ", "NVDA", "TSLA", "AAPL"]


def test_missing_file_returns_base_only(tmp_path):
    syms = watchlist.read_symbols(tmp_path / "does_not_exist.xlsx")
    assert syms == ["$SPX", "SPY", "QQQ"]


def test_empty_column_a_returns_base_only(tmp_path):
    p = tmp_path / "Top 20.xlsx"
    _make_xlsx(p, [])
    syms = watchlist.read_symbols(p)
    assert syms == ["$SPX", "SPY", "QQQ"]


def test_load_error_none_on_success(tmp_path):
    p = tmp_path / "Top 20.xlsx"
    _make_xlsx(p, ["NVDA"])
    watchlist.read_symbols(p)
    assert watchlist.get_load_error() is None


def test_load_error_set_on_missing_file(tmp_path):
    watchlist.read_symbols(tmp_path / "nope.xlsx")
    assert watchlist.get_load_error() is not None


def test_cache_rereads_on_mtime_change(tmp_path, monkeypatch):
    p = tmp_path / "Top 20.xlsx"
    _make_xlsx(p, ["NVDA"])
    monkeypatch.setattr(watchlist, "WATCHLIST_PATH", p)
    watchlist._cache.update(mtime=None, symbols=None)

    first = watchlist.get_scan_symbols()
    assert first == ["$SPX", "SPY", "QQQ", "NVDA"]

    _make_xlsx(p, ["TSLA"])
    st = p.stat()
    os.utime(p, (st.st_atime, st.st_mtime + 10))

    second = watchlist.get_scan_symbols()
    assert second == ["$SPX", "SPY", "QQQ", "TSLA"]

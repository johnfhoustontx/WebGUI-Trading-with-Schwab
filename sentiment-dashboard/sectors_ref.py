"""Sector reference data — workbook loader + S&P cap weights.

Extracted from the (uncopied) tk dashboard so the webgui Sentiment page and
the headless snapshot can build the ``sector_data`` rows that ``scoring``
and ``history_backfill`` consume, without importing tkinter.
"""
from pathlib import Path

SECTORS_XLSX = Path(__file__).parent / "Sectors_Industries_ETFs.xlsx"

# S&P 500 sector weights (percent, ~sum 100) — drives cap-weighted sector_perf.
SP500_SECTOR_WEIGHTS = {
    "Information Technology": 32.53,
    "Financials":             13.42,
    "Communication Services": 10.16,
    "Consumer Discretionary":  9.94,
    "Industrials":             8.86,
    "Health Care":             8.63,
    "Energy":                  4.89,
    "Consumer Staples":        4.61,
    "Materials":               2.74,
    "Real Estate":             2.12,
    "Utilities":               2.09,
}

CYCLICAL_SECTORS = {
    "Consumer Discretionary", "Financials", "Industrials",
    "Information Technology", "Communication Services",
    "Materials", "Energy",
}
DEFENSIVE_SECTORS = {
    "Consumer Staples", "Utilities", "Health Care", "Real Estate",
}


# The workbook is static reference data but load_sectors_data() is called several
# times per sentiment refresh — cache the parsed rows, re-reading only when the
# file mtime changes (mirrors watchlist.get_scan_symbols). Single-entry by path.
_cache = {"path": None, "mtime": None, "data": None}


def reset_cache():
    """Drop the cached rows so the next load re-reads the workbook (test helper)."""
    _cache.update(path=None, mtime=None, data=None)


_stocks_cache = {"path": None, "mtime": None, "data": None}


def reset_stocks_cache():
    """Drop the cached Stocks rows so the next load re-reads (test helper)."""
    _stocks_cache.update(path=None, mtime=None, data=None)


def load_stocks_data(xlsx_path=SECTORS_XLSX):
    """Load the Stocks tab — 5 constituents per (sector, industry) — mtime-cached.

    Returns a list of dicts in workbook order:
        {sector, industry, rank, symbol, company, etfs}
    Returns [] if the workbook, the tab, or openpyxl is unavailable.
    """
    key = str(xlsx_path)
    try:
        mtime = Path(xlsx_path).stat().st_mtime
    except OSError:
        mtime = None
    if (_stocks_cache["data"] is not None and _stocks_cache["path"] == key
            and _stocks_cache["mtime"] == mtime and mtime is not None):
        return _stocks_cache["data"]
    data = _load_stocks_data_uncached(xlsx_path)
    if mtime is not None:
        _stocks_cache.update(path=key, mtime=mtime, data=data)
    return data


def _load_stocks_data_uncached(xlsx_path):
    try:
        import openpyxl
    except ImportError:
        return []
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    except Exception:
        return []
    if "Stocks" not in wb.sheetnames:
        return []

    rows = []
    for i, row in enumerate(wb["Stocks"].iter_rows(values_only=True)):
        if i == 0 or not row:
            continue
        symbol = (row[3] or "") if len(row) > 3 else ""
        symbol = str(symbol).strip()
        # A falsy symbol is the blank line before the trailing merged note
        # block — guard on the data, not on a row number that moves.
        if not symbol:
            continue
        etfs_raw = str(row[5]) if len(row) > 5 and row[5] else ""
        rows.append({
            "sector": (row[0] or "").strip() if row[0] else "",
            "industry": (row[1] or "").strip() if len(row) > 1 and row[1] else "",
            "rank": row[2] if len(row) > 2 else None,
            "symbol": symbol,
            "company": (str(row[4]).strip() if len(row) > 4 and row[4] else ""),
            "etfs": [e.strip() for e in etfs_raw.split(",") if e.strip()],
        })
    return rows


def stock_symbols(xlsx_path=SECTORS_XLSX):
    """Deduped constituent symbols in workbook order.

    A name legitimately appears under more than one industry, so the fetch
    universe is the deduped list, not the row count.
    """
    seen, out = set(), []
    for row in load_stocks_data(xlsx_path):
        if row["symbol"] not in seen:
            seen.add(row["symbol"])
            out.append(row["symbol"])
    return out


def constituents_by_industry(xlsx_path=SECTORS_XLSX):
    """{(sector, industry): [symbols]} — the participation input."""
    out = {}
    for row in load_stocks_data(xlsx_path):
        out.setdefault((row["sector"], row["industry"]), []).append(row["symbol"])
    return out


def load_sectors_data(xlsx_path=SECTORS_XLSX):
    """Load sector / industry / ETF rows from the reference workbook (mtime-cached).

    Returns a list of dicts in display order:
        {kind: 'sector'|'industry', sector, label, etf, name, notes, sp_weight}
    Returns [] if the workbook or openpyxl is unavailable.
    """
    key = str(xlsx_path)
    try:
        mtime = Path(xlsx_path).stat().st_mtime
    except OSError:
        mtime = None
    if (_cache["data"] is not None and _cache["path"] == key
            and _cache["mtime"] == mtime and mtime is not None):
        return _cache["data"]
    data = _load_sectors_data_uncached(xlsx_path)
    if mtime is not None:  # only cache successful, stat-able loads
        _cache.update(path=key, mtime=mtime, data=data)
    return data


def _load_sectors_data_uncached(xlsx_path):
    try:
        import openpyxl
    except ImportError:
        return []
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    except Exception:
        return []

    sector_primary = {}
    sector_descriptions = {}
    sector_order = []
    for i, row in enumerate(wb["Sectors"].iter_rows(values_only=True)):
        if i == 0 or not row or not row[1]:
            continue
        sector_name, spdr = row[1], row[2]
        description = row[5] if len(row) > 5 else None
        sector_primary[sector_name] = spdr
        sector_descriptions[sector_name] = description or ''
        sector_order.append(sector_name)

    industries_by_sector = {}
    seen_industry = set()  # (sector, industry) — keep first ETF only
    seen_etf = set()       # global ETF dedupe (e.g. MJ in Staples + Health)
    for i, row in enumerate(wb["Industries"].iter_rows(values_only=True)):
        if i == 0 or not row or not row[0]:
            continue
        sector = row[0]
        industry = row[1] if len(row) > 1 else None
        etf = row[2] if len(row) > 2 else None
        etf_name = row[3] if len(row) > 3 else None
        notes = row[5] if len(row) > 5 else None
        if not etf:
            continue
        key = (sector, industry)
        if key in seen_industry or etf in seen_etf:
            continue
        seen_industry.add(key)
        seen_etf.add(etf)
        industries_by_sector.setdefault(sector, []).append(
            {'industry': industry, 'etf': etf,
             'name': etf_name or '', 'notes': notes or ''})

    rows = []
    for sector_name in sector_order:
        rows.append({
            'kind': 'sector', 'sector': sector_name,
            'label': sector_name, 'etf': sector_primary.get(sector_name),
            'name': sector_descriptions.get(sector_name, ''), 'notes': '',
            'sp_weight': SP500_SECTOR_WEIGHTS.get(sector_name, 0.0),
        })
        for ind in industries_by_sector.get(sector_name, []):
            rows.append({
                'kind': 'industry', 'sector': sector_name,
                'label': ind['industry'] or ind['etf'], 'etf': ind['etf'],
                'name': ind['name'], 'notes': ind['notes'],
                'sp_weight': 0.0,
            })
    return rows

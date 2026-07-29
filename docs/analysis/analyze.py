"""Intraday correlation / relationship analysis of the watchlist vs indexes and
other indicators. Reads prices.pkl (5-min RTH candles) and writes:
  docs/analysis/2026-07-21-intraday-correlation.xlsx
  docs/analysis/2026-07-21-correlation-heatmap.png
  docs/analysis/2026-07-21-stockstock-clustermap.png
  docs/analysis/2026-07-21-stock-index-correlation.md
"""
import pickle, numpy as np, pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from scipy.cluster.hierarchy import linkage, leaves_list
from scipy.spatial.distance import squareform

HERE = r"C:/Users/john_/AppData/Local/Temp/claude/D--WebGUI-Trading-with-Schwab/9a392989-bcd2-4e24-937f-2bd010329c64/scratchpad"
OUT  = r"D:/WebGUI Trading with Schwab/docs/analysis"
DATE = "2026-07-21"

with open(f"{HERE}/prices.pkl","rb") as f:
    blob = pickle.load(f)
P, G = blob["prices"], blob["groups"]
STOCKS, BROAD, INDEX, VOL, BREADTH, SECTORS = (G["STOCKS"],G["BROAD_ETF"],G["INDEX"],
                                               G["VOL"],G["BREADTH"],G["SECTORS"])
ANALYSED = STOCKS + BROAD                     # names we characterise (have volume)

# ---- within-session 5-min log returns -------------------------------------
def sess(df): return df.index.normalize()
def logret_within(close):
    lr = np.log(close/close.shift(1))
    lr[sess(close.to_frame()).to_series().values != sess(close.to_frame()).shift(1).to_series().values] = np.nan
    return lr
# simpler/robust session-aware return
def wret(df):
    c = df["close"]; g = c.groupby(c.index.date)
    return g.apply(lambda s: np.log(s/s.shift(1))).reset_index(level=0, drop=True)

ret = pd.DataFrame({s: wret(P[s]) for s in P if s not in BREADTH})
# breadth: net advancers delta within session
adv, dec = P["$ADVN"]["close"], P["$DECN"]["close"]
adnet = (adv - dec)
adnet_delta = adnet.groupby(adnet.index.date).apply(lambda s: s.diff()).reset_index(level=0, drop=True)
ret["AD_BREADTH"] = adnet_delta

R = ret.dropna(how="any")          # all symbols share the session-first NaN rows
N = len(R)

REFS = ["SPY","QQQ","$SPX","$NDX","IWM","DIA","$VIX","AD_BREADTH"] + SECTORS
CORR = R.corr()

# ---- per-name summary ------------------------------------------------------
def beta_r2(y, x):
    x = R[x]; y = R[y]
    b = np.cov(y, x)[0,1] / np.var(x)
    r = np.corrcoef(y, x)[0,1]
    return b, r*r

def obv_trend(df):
    d = np.sign(df["close"].diff().fillna(0)) * df["volume"]
    obv = d.cumsum().values
    t = np.arange(len(obv))
    slope = np.polyfit(t, obv, 1)[0]
    denom = df["volume"].mean() or 1
    norm = slope / denom                       # bars of net vol per bar
    return norm

def macd_state(close):
    e12 = close.ewm(span=12, adjust=False).mean()
    e26 = close.ewm(span=26, adjust=False).mean()
    macd = e12 - e26
    sig = macd.ewm(span=9, adjust=False).mean()
    hist = macd.iloc[-1] - sig.iloc[-1]
    return ("Bull" if macd.iloc[-1] > sig.iloc[-1] else "Bear"), float(hist)

def vwap_pct_last(df):
    last_day = df.index.date.max()
    d = df[df.index.date == last_day]
    tp = (d["high"]+d["low"]+d["close"])/3
    vwap = (tp*d["volume"]).cumsum()/d["volume"].cumsum().replace(0,np.nan)
    above = (d["close"] > vwap).mean()*100
    return float(above)

def rel_strength(y):
    rs = np.exp((R[y] - R["SPY"]).cumsum())     # RS line vs SPY (starts ~1)
    final = rs.iloc[-1]
    slope = np.polyfit(np.arange(len(rs)), rs.values, 1)[0]*len(rs)  # total drift
    return float(final), float(slope)

rows = []
for s in ANALYSED:
    b_spy, r2_spy = beta_r2(s, "SPY")
    spx_c = CORR.loc[s, "$SPX"]; ndx_c = CORR.loc[s, "$NDX"]
    sect_c = CORR.loc[s, SECTORS]
    top_sect = sect_c.idxmax(); top_sect_c = sect_c.max()
    vix_c = CORR.loc[s, "$VIX"]
    ad_c = CORR.loc[s, "AD_BREADTH"]
    mkt_avg = CORR.loc[s, [x for x in ANALYSED if x != s]].mean()
    obv = obv_trend(P[s])
    macd_dir, macd_hist = macd_state(P[s]["close"])
    vwap_ab = vwap_pct_last(P[s])
    rs_final, rs_slope = rel_strength(s)
    rows.append(dict(Symbol=s, Beta_SPY=b_spy, R2_SPY=r2_spy,
                     Corr_SPX=spx_c, Corr_NDX=ndx_c,
                     Index_lean=("NDX" if ndx_c>spx_c else "SPX"),
                     Top_Sector=top_sect, Sector_Corr=top_sect_c,
                     Corr_VIX=vix_c, Corr_Breadth=ad_c,
                     Avg_Corr_Watchlist=mkt_avg,
                     RS_vs_SPY=rs_final, OBV_trend=obv,
                     MACD=macd_dir, MACD_hist=macd_hist, Pct_above_VWAP=vwap_ab))
S = pd.DataFrame(rows).set_index("Symbol")

# ---- correlation matrix (analysed rows x references) ----------------------
MAT = CORR.loc[ANALYSED, REFS].copy()

# ---- stock-stock clustering -----------------------------------------------
SS = CORR.loc[STOCKS, STOCKS].copy()
dist = 1 - SS.values
np.fill_diagonal(dist, 0.0)
dist = (dist + dist.T)/2
Z = linkage(squareform(dist, checks=False), method="average")
order = leaves_list(Z)
SS_ord = SS.iloc[order, order]

# =================== WRITE EXCEL ===========================================
xlsx = f"{OUT}/{DATE}-intraday-correlation.xlsx"
with pd.ExcelWriter(xlsx, engine="openpyxl") as xw:
    MAT.round(3).to_excel(xw, sheet_name="Corr_Matrix")
    S.round(3).to_excel(xw, sheet_name="Summary")
    CORR.loc[REFS, REFS].round(3).to_excel(xw, sheet_name="Reference_Grid")
    SS_ord.round(3).to_excel(xw, sheet_name="Stock_Stock_Clustered")
    cov = pd.DataFrame({"Symbol":list(P.keys()),
                        "bars":[len(P[s]) for s in P],
                        "first":[str(P[s].index[0]) for s in P],
                        "last":[str(P[s].index[-1]) for s in P]})
    cov.to_excel(xw, sheet_name="Coverage", index=False)

# conditional formatting (blue-white-red color scale) on matrix-like sheets
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
wb = load_workbook(xlsx)
scale = lambda: ColorScaleRule(start_type="num", start_value=-1, start_color="F8696B",
                               mid_type="num", mid_value=0, mid_color="FFFFFF",
                               end_type="num", end_value=1, end_color="4F9BFF")
for sh, ncol, nrow in [("Corr_Matrix", len(REFS), len(ANALYSED)),
                       ("Reference_Grid", len(REFS), len(REFS)),
                       ("Stock_Stock_Clustered", len(STOCKS), len(STOCKS))]:
    ws = wb[sh]
    last_col = chr(ord('A')+ncol)  # ok for <=25 cols; matrix cols may exceed -> use util
    from openpyxl.utils import get_column_letter
    rng = f"B2:{get_column_letter(ncol+1)}{nrow+1}"
    ws.conditional_formatting.add(rng, scale())
    ws.freeze_panes = "B2"
wb.save(xlsx)
print("wrote", xlsx)

# =================== HEATMAP PNGs ==========================================
def heat(df, title, path, figsize):
    fig, ax = plt.subplots(figsize=figsize)
    im = ax.imshow(df.values, cmap="RdBu_r", vmin=-1, vmax=1, aspect="auto")
    ax.set_xticks(range(len(df.columns))); ax.set_xticklabels(df.columns, rotation=90, fontsize=7)
    ax.set_yticks(range(len(df.index))); ax.set_yticklabels(df.index, fontsize=7)
    ax.set_title(title, fontsize=11)
    fig.colorbar(im, ax=ax, fraction=0.025, pad=0.02)
    fig.tight_layout(); fig.savefig(path, dpi=140); plt.close(fig)
    print("wrote", path)

# order analysed rows by their SPY beta for readability
mat_ord = MAT.loc[S.sort_values("Beta_SPY", ascending=False).index]
heat(mat_ord, f"Intraday 5-min return correlation — watchlist vs references ({DATE})",
     f"{OUT}/{DATE}-correlation-heatmap.png", (11, 14))
heat(SS_ord, f"Stock-stock intraday correlation (clustered) ({DATE})",
     f"{OUT}/{DATE}-stockstock-clustermap.png", (13, 12))

# save computed frames for the report step
S.to_pickle(f"{HERE}/summary.pkl")
MAT.to_pickle(f"{HERE}/matrix.pkl")
CORR.to_pickle(f"{HERE}/corr_full.pkl")
pd.Series({"N":N,"start":str(R.index[0]),"end":str(R.index[-1]),
           "sessions":R.index.normalize().nunique()}).to_pickle(f"{HERE}/meta.pkl")
print("N return obs:", N, "sessions:", R.index.normalize().nunique())
print("\n=== quick sanity ===")
print("SPY-QQQ corr:", round(CORR.loc['SPY','QQQ'],3), " (expect high ~.9)")
print("SPY-$VIX corr:", round(CORR.loc['SPY','$VIX'],3), " (expect strongly negative)")
print("mean beta:", round(S.Beta_SPY.mean(),2))
print("highest beta:\n", S.Beta_SPY.sort_values(ascending=False).head(6).round(2).to_string())
print("lowest R2 (idiosyncratic):\n", S.R2_SPY.sort_values().head(6).round(2).to_string())

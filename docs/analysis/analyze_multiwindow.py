"""Multi-window DAILY correlation pass (1mo/3mo/6mo) + stability vs the intraday result.
Appends sheets to the existing workbook, writes a stability heatmap, and prints the
tables used for the report addendum."""
import pickle, numpy as np, pandas as pd
import matplotlib; matplotlib.use("Agg")
import matplotlib.pyplot as plt

HERE = r"C:/Users/john_/AppData/Local/Temp/claude/D--WebGUI-Trading-with-Schwab/9a392989-bcd2-4e24-937f-2bd010329c64/scratchpad"
OUT  = r"D:/WebGUI Trading with Schwab/docs/analysis"
DATE = "2026-07-21"
XLSX = f"{OUT}/{DATE}-intraday-correlation.xlsx"

d = pickle.load(open(f"{HERE}/daily.pkl","rb"))
P, G = d["prices"], d["groups"]
STOCKS, BROAD, INDEX, VOL, BREADTH, SECTORS = (G["STOCKS"],G["BROAD_ETF"],G["INDEX"],
                                               G["VOL"],G["BREADTH"],G["SECTORS"])
ANALYSED = STOCKS + BROAD
REFS = ["SPY","QQQ","$SPX","$NDX","IWM","DIA","$VIX","AD_BREADTH"] + SECTORS

# ---- daily returns ---------------------------------------------------------
ret = {}
for s in P:
    if s in BREADTH: continue
    c = P[s]["close"]
    ret[s] = np.log(c/c.shift(1))
adnet = (P["$ADVN"]["close"] - P["$DECN"]["close"])
ret["AD_BREADTH"] = adnet.diff()
RET = pd.DataFrame(ret).sort_index()
RET = RET.iloc[1:]                      # drop first NaN row

WINDOWS = {"1mo":21, "3mo":63, "6mo":126}

def pair(R, y, x):
    dd = R[[y,x]].replace([np.inf,-np.inf],np.nan).dropna()
    if len(dd) < 8: return np.nan, np.nan, np.nan, len(dd)
    rho = np.corrcoef(dd[y],dd[x])[0,1]
    beta = np.cov(dd[y],dd[x])[0,1]/np.var(dd[x])
    return rho, beta, rho*rho, len(dd)

CORRW, MATW = {}, {}
for name,n in WINDOWS.items():
    Rw = RET.tail(n)
    CORRW[name] = Rw.corr()
    MATW[name]  = CORRW[name].loc[ANALYSED, REFS]

# intraday reference (from the first analysis)
CI = pd.read_pickle(f"{HERE}/corr_full.pkl")     # intraday corr
SI = pd.read_pickle(f"{HERE}/summary.pkl")        # intraday summary (Top_Sector etc.)

# ---- stability table -------------------------------------------------------
rows=[]
for s in ANALYSED:
    rho_i = CI.loc[s,"SPY"]
    rho = {w: CORRW[w].loc[s,"SPY"] for w in WINDOWS}
    beta = {w: pair(RET.tail(WINDOWS[w]), s, "SPY")[1] for w in WINDOWS}
    n3 = pair(RET.tail(63), s, "SPY")[3]
    # daily 3mo dominant sector
    sec3 = CORRW["3mo"].loc[s, SECTORS]
    top3 = sec3.idxmax(); top3c = sec3.max()
    sec_intr = SI.loc[s,"Top_Sector"]
    daily_rhos = [rho["1mo"],rho["3mo"],rho["6mo"]]
    drift = np.nanmax(daily_rhos)-np.nanmin(daily_rhos)
    # sign consistency intraday vs daily-3mo
    flip = "" if np.sign(rho_i)==np.sign(rho["3mo"]) else "SIGN-FLIP"
    lean3 = "NDX" if CORRW["3mo"].loc[s,"$NDX"]>CORRW["3mo"].loc[s,"$SPX"] else "SPX"
    rows.append(dict(Symbol=s, N_3mo=n3,
                     rSPY_intraday=rho_i, rSPY_1mo=rho["1mo"], rSPY_3mo=rho["3mo"], rSPY_6mo=rho["6mo"],
                     rSPY_drift=drift, bSPY_1mo=beta["1mo"], bSPY_3mo=beta["3mo"], bSPY_6mo=beta["6mo"],
                     Lean_3mo=lean3, Sector_intraday=sec_intr, Sector_3mo=top3, Sector3_r=top3c,
                     Sector_stable=("yes" if sec_intr==top3 else "no"), Flag=flip))
STAB = pd.DataFrame(rows).set_index("Symbol")

# ---- append to workbook ----------------------------------------------------
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
with pd.ExcelWriter(XLSX, mode="a", engine="openpyxl", if_sheet_exists="replace") as xw:
    for w in WINDOWS: MATW[w].round(3).to_excel(xw, sheet_name=f"Daily_{w}")
    STAB.round(3).to_excel(xw, sheet_name="Stability")
wb = load_workbook(XLSX)
scale = lambda: ColorScaleRule(start_type="num",start_value=-1,start_color="F8696B",
                               mid_type="num",mid_value=0,mid_color="FFFFFF",
                               end_type="num",end_value=1,end_color="4F9BFF")
for w in WINDOWS:
    ws=wb[f"Daily_{w}"]; ws.freeze_panes="B2"
    ws.conditional_formatting.add(f"B2:{get_column_letter(len(REFS)+1)}{len(ANALYSED)+1}", scale())
ws=wb["Stability"]; ws.freeze_panes="B2"
# color the 4 rSPY columns (B..E) and drift
ws.conditional_formatting.add(f"C2:F{len(ANALYSED)+1}", scale())
wb.save(XLSX)
print("appended Daily_1mo/3mo/6mo + Stability to", XLSX)

# ---- stability heatmap -----------------------------------------------------
H = STAB[["rSPY_intraday","rSPY_1mo","rSPY_3mo","rSPY_6mo"]].copy()
H.columns=["intraday 5m","daily 1mo","daily 3mo","daily 6mo"]
H = H.loc[SI.sort_values("Beta_SPY",ascending=False).index.intersection(H.index).tolist()
          + [x for x in H.index if x not in SI.index]]
H = H.reindex(SI.sort_values("Beta_SPY",ascending=False).index)
fig,ax=plt.subplots(figsize=(6,14))
im=ax.imshow(H.values,cmap="RdBu_r",vmin=-1,vmax=1,aspect="auto")
ax.set_xticks(range(4)); ax.set_xticklabels(H.columns,rotation=30,ha="right",fontsize=8)
ax.set_yticks(range(len(H))); ax.set_yticklabels(H.index,fontsize=7)
for i in range(len(H)):
    for j in range(4):
        v=H.values[i,j]
        if pd.notna(v): ax.text(j,i,f"{v:.2f}",ha="center",va="center",fontsize=6,
                                color="white" if abs(v)>0.55 else "black")
ax.set_title("Correlation with SPY across horizons\n(stability of the market relationship)",fontsize=10)
fig.colorbar(im,ax=ax,fraction=0.04,pad=0.03)
fig.tight_layout(); fig.savefig(f"{OUT}/{DATE}-stability-across-horizons.png",dpi=140); plt.close(fig)
print("wrote", f"{OUT}/{DATE}-stability-across-horizons.png")

STAB.to_pickle(f"{HERE}/stability.pkl")

# ---- printouts for the report ---------------------------------------------
pd.set_option("display.width",220,"display.max_columns",30,"display.max_rows",60)
print("\n===== STABILITY (sorted by daily 3mo rho SPY desc) =====")
show=["N_3mo","rSPY_intraday","rSPY_1mo","rSPY_3mo","rSPY_6mo","rSPY_drift","bSPY_3mo","bSPY_6mo","Lean_3mo","Sector_intraday","Sector_3mo","Sector_stable","Flag"]
print(STAB[show].sort_values("rSPY_3mo",ascending=False).round(2).to_string())

print("\n===== reference cross-corr, daily 6mo =====")
print(CORRW["6mo"].loc[["SPY","QQQ","$NDX","IWM","DIA","$VIX","AD_BREADTH"],
                       ["SPY","QQQ","$NDX","IWM","$VIX","AD_BREADTH"]].round(2).to_string())

print("\n===== HEADLINE stability stats =====")
sec_stab = (STAB["Sector_stable"]=="yes").mean()*100
flips = STAB.index[STAB["Flag"]=="SIGN-FLIP"].tolist()
print(f"sector mapping intraday==daily3mo: {sec_stab:.0f}% of names")
print("SIGN-FLIP (intraday vs daily-3mo market corr):", flips)
print("mean |drift| daily 1->6mo:", round(STAB.rSPY_drift.mean(),2))
print("names most UNstable (drift):"); print(STAB.rSPY_drift.sort_values(ascending=False).head(6).round(2).to_string())
print("defensives daily-3mo rSPY:")
print(STAB.loc[[x for x in ["WMT","T","XOM","ABBV","CMCSA","PFE"] if x in STAB.index],
               ["rSPY_intraday","rSPY_3mo","rSPY_6mo"]].round(2).to_string())

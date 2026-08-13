"""Per-symbol 1-8 week swing signal + detailed technical/relationship card.

DUAL-LENS by design:
  (1) FACTOR MODEL — replicates the app's VALIDATED 20-trading-day model
      (trade-analyzer/data/swing_model.json): same factors, signed IC-weights,
      cross-sectionally z-scored across the current equity universe, mapped through the
      model's own calibration bands. A mean-reversion / momentum / vol-premium view.
  (2) TECHNICAL POSTURE — classical trend/momentum (Wilder's EMA stack, MACD, RSI, ADX).
An AGREEMENT flag marks where the two lenses confirm vs conflict (conflict = low confidence).
Index ETFs are NOT scored by the single-name factor model (basket => n/a), only posture.

SYSTEMATIC MODEL OUTPUT — NOT personalized financial advice.
"""
import json, pickle, numpy as np, pandas as pd

HERE = r"C:/Users/john_/AppData/Local/Temp/claude/D--WebGUI-Trading-with-Schwab/9a392989-bcd2-4e24-937f-2bd010329c64/scratchpad"
OUT  = r"D:/WebGUI Trading with Schwab/docs/analysis"
DATE = "2026-07-21"
XLSX = f"{OUT}/{DATE}-intraday-correlation.xlsx"

d = pickle.load(open(f"{HERE}/daily2y.pkl","rb"))
P, G = d["prices"], d["groups"]
STOCKS, BROAD, SECTORS = G["STOCKS"], G["BROAD_ETF"], G["SECTORS"]
UNIV = STOCKS + BROAD
MODEL = json.load(open(r"D:/WebGUI Trading with Schwab/trade-analyzer/data/swing_model.json"))
REG = MODEL["regimes"]["all"]; W = REG["weights"]; CAL = REG["calibration"]
STAB = pd.read_pickle(f"{HERE}/stability.pkl")
SUMI = pd.read_pickle(f"{HERE}/summary.pkl")

SECMAP = {s: STAB.loc[s,"Sector_3mo"] if s in STAB.index else "SPY" for s in UNIV}
SECMAP.update({"SPY":"SPY","QQQ":"XLK","IWM":"XLI","DIA":"XLI"})

def rma(s,n): return s.ewm(alpha=1/n, adjust=False, min_periods=n).mean()
def rsi(c,n=14):
    dl=c.diff(); g=dl.clip(lower=0); l=-dl.clip(upper=0)
    rs=rma(g,n)/rma(l,n); return (100-100/(1+rs)).iloc[-1]
def atr_val(h,l,c,n=14):
    pc=c.shift(1); tr=pd.concat([h-l,(h-pc).abs(),(l-pc).abs()],axis=1).max(axis=1)
    a=rma(tr,n); return (a/c*100).iloc[-1], a.iloc[-1]
def adx(h,l,c,n=14):
    up=h.diff(); dn=-l.diff()
    pdm=np.where((up>dn)&(up>0),up,0.0); mdm=np.where((dn>up)&(dn>0),dn,0.0)
    pc=c.shift(1); tr=pd.concat([h-l,(h-pc).abs(),(l-pc).abs()],axis=1).max(axis=1)
    a=rma(tr,n); pdi=100*rma(pd.Series(pdm,index=c.index),n)/a
    mdi=100*rma(pd.Series(mdm,index=c.index),n)/a
    dx=100*(pdi-mdi).abs()/(pdi+mdi).replace(0,np.nan); return rma(dx,n).iloc[-1]
def macd_state(c):
    m=c.ewm(span=12,adjust=False).mean()-c.ewm(span=26,adjust=False).mean()
    sig=m.ewm(span=9,adjust=False).mean(); h=m-sig
    return ("Bullish" if m.iloc[-1]>sig.iloc[-1] else "Bearish"), float(h.iloc[-1]), float(h.iloc[-1]-h.iloc[-2])
def ema(c,n): return c.ewm(span=n,adjust=False).mean().iloc[-1]

def factors(sym):
    c=P[sym]["close"]; v=P[sym]["volume"]; n=len(c); r=np.log(c/c.shift(1)); f={}
    f["mom_12_1"]=c.iloc[-21]/c.iloc[-273]-1 if n>=273 else np.nan
    f["mom_6_1"] =c.iloc[-21]/c.iloc[-126]-1 if n>=126 else np.nan
    rv=r.iloc[-60:].std()*np.sqrt(252) if n>=60 else np.nan
    f["low_vol"]=-rv
    e50=c.ewm(span=50,adjust=False).mean().iloc[-1]; e200=c.ewm(span=200,adjust=False).mean().iloc[-1] if n>=200 else np.nan
    f["trend_quality"]=0.5*(c.iloc[-1]/e50-1)+0.5*(e50/e200-1) if e200==e200 else (c.iloc[-1]/e50-1)
    sec=P[SECMAP[sym]]["close"]
    f["rs_sector"]=(c.iloc[-1]/c.iloc[-64]-1)-(sec.iloc[-1]/sec.iloc[-64]-1) if n>=64 and len(sec)>=64 else np.nan
    f["turnover"]=v.iloc[-5:].mean()/v.iloc[-63:].mean() if n>=63 and v.iloc[-63:].mean()>0 else np.nan
    return f

F=pd.DataFrame({s:factors(s) for s in UNIV}).T[list(W.keys())]
def zsc(col):
    base=col.loc[STOCKS].dropna(); lo,hi=np.percentile(base,2),np.percentile(base,98)
    mu,sd=base.clip(lo,hi).mean(),base.clip(lo,hi).std()
    return ((col.clip(lo,hi)-mu)/sd if sd>0 else col*0).clip(-3,3)
Z=F.apply(zsc)
comp=(Z.fillna(0)*pd.Series(W)).sum(axis=1)
def band_of(x):
    for b in CAL:
        if b["score_lo"]<=x<=b["score_hi"]: return b
    return CAL[0] if x<CAL[0]["score_lo"] else CAL[-1]
def fsignal(x):
    b=band_of(x); return ("BUY" if b["band"]==4 else "SELL" if b["band"]==0 else "HOLD"), b
pct=comp.loc[STOCKS].rank(pct=True)*100   # percentile among stocks only

# ---- technical posture -----------------------------------------------------
def posture_score(trend,macd,mslope,rsi_,adx_,chg1m,px,e200):
    s=0.0
    s+={"Bull":2,"Mixed":0,"Bear":-2}[trend]
    s+=1 if macd=="Bullish" else -1
    s+=0.5 if mslope>0 else -0.5
    s+=1 if rsi_>=55 else (-1 if rsi_<=45 else 0)
    if rsi_>=75: s-=1.0                       # overbought => pullback risk
    if rsi_<=25: s+=1.0                       # oversold => bounce potential
    if not np.isnan(e200): s+=0.5 if px>e200 else -0.5
    s+=0.5*np.sign(chg1m)
    if adx_>=25: s+=0.5*np.sign(s)            # strong trend amplifies direction
    return s
def posture_label(s):
    return "Bullish" if s>=2 else "Bearish" if s<=-2 else "Neutral"

rows=[]
for s in UNIV:
    c=P[s]["close"]; h=P[s]["high"]; l=P[s]["low"]; n=len(c); px=c.iloc[-1]
    chg=lambda k:(px/c.iloc[-k]-1)*100 if n>k else np.nan
    e20,e50=ema(c,20),ema(c,50); e200=ema(c,200) if n>=200 else np.nan
    win=c.iloc[-252:] if n>=252 else c
    pos52=(px-win.min())/(win.max()-win.min())*100 if win.max()>win.min() else np.nan
    stack=("Bull" if px>e20>e50 and (np.isnan(e200) or e50>e200) else
           "Bear" if px<e20<e50 and (np.isnan(e200) or e50<e200) else "Mixed")
    _rsi=rsi(c); ap,atr=atr_val(h,l,c); _adx=adx(h,l,c); md,mh,mslope=macd_state(c)
    ext=(px-e20)/atr if atr>0 else np.nan
    c1m=chg(21)
    pscore=posture_score(stack,md,mslope,_rsi,_adx,0 if np.isnan(c1m) else c1m,px,e200)
    plabel=posture_label(pscore)
    is_etf=s in BROAD
    fsig,b=("n/a",band_of(comp[s])) if is_etf else fsignal(comp[s])
    # agreement (only where factor model applies)
    fdir={"BUY":1,"SELL":-1,"HOLD":0,"n/a":None}[fsig]
    pdir={"Bullish":1,"Bearish":-1,"Neutral":0}[plabel]
    if fdir is None: agree="ETF — posture only"
    elif fdir==0 or pdir==0: agree="Mixed"
    elif fdir==pdir: agree=("Confirmed bull" if fdir>0 else "Confirmed bear")
    else: agree="CONFLICT (low conf.)"
    rows.append(dict(Symbol=s, Factor=fsig, Posture=plabel, Agreement=agree,
        Pctile=(np.nan if is_etf else pct[s]), Composite=comp[s], ExpFwd20d=b["mean_fwd"]*100,
        HitRate=b["hit_rate"]*100, Band=b["band"], Price=px, Chg_1w=chg(5), Chg_1m=c1m,
        Chg_3m=chg(63), Pos_52wk=pos52, Trend=stack, RSI14=_rsi, RSI_ob=("OB" if _rsi>=70 else "OS" if _rsi<=30 else ""),
        ADX14=_adx, MACD=md, MACD_hist=mh, MACD_slope=mslope, ATR_pct=ap, Ext_ATR=ext,
        DomSector=SECMAP[s], Beta3mo=STAB.loc[s,"bSPY_3mo"] if s in STAB.index else np.nan,
        R2_intraday=SUMI.loc[s,"R2_SPY"] if s in SUMI.index else np.nan,
        RS_SPY_3mo=STAB.loc[s,"rSPY_3mo"] if s in STAB.index else np.nan,
        VIXcorr=SUMI.loc[s,"Corr_VIX"] if s in SUMI.index else np.nan,
        z_mom12=Z.loc[s,"mom_12_1"], z_mom6=Z.loc[s,"mom_6_1"], z_lowvol=Z.loc[s,"low_vol"],
        z_trend=Z.loc[s,"trend_quality"], z_rss=Z.loc[s,"rs_sector"], z_turn=Z.loc[s,"turnover"],
        Young=("*" if n<273 else "")))
T=pd.DataFrame(rows).set_index("Symbol")
# order: stocks by composite desc, ETFs last
T=pd.concat([T.loc[[s for s in T.index if s in STOCKS]].sort_values("Composite",ascending=False),
             T.loc[[s for s in T.index if s in BROAD]]])

# ---- Excel -----------------------------------------------------------------
xcols=["Factor","Posture","Agreement","Pctile","Composite","ExpFwd20d","HitRate","Price","Chg_1w",
       "Chg_1m","Chg_3m","Pos_52wk","Trend","RSI14","ADX14","MACD","ATR_pct","Ext_ATR","DomSector",
       "Beta3mo","RS_SPY_3mo","VIXcorr","z_mom12","z_mom6","z_lowvol","z_trend","z_rss","z_turn","Young"]
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
with pd.ExcelWriter(XLSX, mode="a", engine="openpyxl", if_sheet_exists="replace") as xw:
    T[xcols].round(3).to_excel(xw, sheet_name="Swing_Signals")
wb=load_workbook(XLSX); ws=wb["Swing_Signals"]; ws.freeze_panes="B2"
ff={"BUY":"C6EFCE","HOLD":"FFEB9C","SELL":"FFC7CE"}; fn={"BUY":"006100","HOLD":"9C6500","SELL":"9C0006"}
pf={"Bullish":"C6EFCE","Neutral":"FFEB9C","Bearish":"FFC7CE"}
for r in range(2,len(T)+2):
    fv=ws.cell(r,2).value; pv=ws.cell(r,3).value; av=ws.cell(r,4).value
    if fv in ff: ws.cell(r,2).fill=PatternFill("solid",fgColor=ff[fv]); ws.cell(r,2).font=Font(color=fn[fv],bold=True)
    if pv in pf: ws.cell(r,3).fill=PatternFill("solid",fgColor=pf[pv])
    if isinstance(av,str) and av.startswith("CONFLICT"): ws.cell(r,4).fill=PatternFill("solid",fgColor="FCE4D6"); ws.cell(r,4).font=Font(color="9C0006")
wb.save(XLSX); print("wrote Swing_Signals sheet")

# ---- per-symbol cards ------------------------------------------------------
NAMES={"AAL":"American Airlines","AAPL":"Apple","ABBV":"AbbVie","AMD":"AMD","AMZN":"Amazon",
"AVGO":"Broadcom","BAC":"Bank of America","CMCSA":"Comcast","CMG":"Chipotle","CRWV":"CoreWeave",
"DELL":"Dell","GOOGL":"Alphabet","HOOD":"Robinhood","INTC":"Intel","IONQ":"IonQ","IREN":"IREN",
"JPM":"JPMorgan","META":"Meta","MRVL":"Marvell","MU":"Micron","NFLX":"Netflix","NVDA":"Nvidia",
"PFE":"Pfizer","PLTR":"Palantir","QCOM":"Qualcomm","RGTI":"Rigetti","RKLB":"Rocket Lab",
"SMCI":"Super Micro","SOFI":"SoFi","SPCX":"SPAC/SPCX","T":"AT&T","TSLA":"Tesla","UBER":"Uber",
"WMT":"Walmart","WULF":"TeraWulf","XOM":"Exxon","ALAB":"Astera Labs","NBIS":"Nebius","MSFT":"Microsoft",
"SPY":"S&P 500 ETF","QQQ":"Nasdaq-100 ETF","IWM":"Russell 2000 ETF","DIA":"Dow 30 ETF"}
def sgn(z):
    if np.isnan(z): return "n/a"
    return "strong+" if z>=1 else "+" if z>=0.3 else "strong-" if z<=-1 else "-" if z<=-0.3 else "~0"
def tech_read(r):
    p=[{"Bull":"Uptrend (price above rising EMAs)","Bear":"Downtrend (price below falling EMAs)",
        "Mixed":"No clean trend (EMAs entangled)"}[r.Trend],
       f"{'strong' if r.ADX14>=25 else 'weak'} trend (ADX {r.ADX14:.0f})",
       f"MACD {r.MACD.lower()}"+(" & improving" if r.MACD_slope>0 else " & weakening"),
       f"RSI {r.RSI14:.0f}"+(" (overbought)" if r.RSI_ob=="OB" else " (oversold)" if r.RSI_ob=="OS" else "")]
    if abs(r.Ext_ATR)>=2.5:
        p.append(f"{'extended' if r.Ext_ATR>0 else 'stretched'} {r.Ext_ATR:+.1f} ATR vs EMA20 — {'pullback' if r.Ext_ATR>0 else 'bounce'} risk")
    return "; ".join(p)+"."

md=[f"# Per-symbol swing signals (1–8 weeks) — {DATE}\n",
"> ### ⚠ SYSTEMATIC MODEL OUTPUT — NOT financial advice\n"
"> These are transparent, rules-based outputs of mechanical models over price data (as of the 2026-07-20 "
"close). They do **not** account for your financial situation, objectives, or risk tolerance and are **not** "
"a recommendation to buy or sell any security. I am not a licensed financial advisor. Signals are "
"horizon-specific (≈1–8 weeks), decay quickly around earnings and macro events, and must be combined with your "
"own diligence, position sizing, and risk management.\n",
"**Two independent lenses are shown per name — read them together:**\n",
"1. **Factor model** — the app's *validated* 20-trading-day model (`swing_model.json`, OOS IC ≈ +0.037, "
"thin/regime-dependent). Cross-sectional over the 39 stocks; BUY = top calibration band (hist. **+1.35%** avg "
"20-day excess return, 52% hit), SELL = bottom band (**−0.80%**, 43% hit), HOLD = middle. It is a "
"**momentum + volatility-premium + mean-reversion** model, *not* a trend follower.\n",
"2. **Technical posture** — classical trend/momentum (Wilder's EMA-20/50/200 stack, MACD, RSI-14, ADX-14).\n",
"**Agreement** flags where they confirm vs **CONFLICT** (opposite → treat as low-confidence). Index ETFs "
"(SPY/QQQ/IWM/DIA) get posture only — a single-name cross-sectional factor model is invalid for a basket.\n",
"> **Why so many factor-model BUYs are beaten-down high-vol names:** the model's largest weight is "
"`low_vol −0.34` (high-volatility names scored *positively* in the bull-ish fit period) plus 6–12-month "
"momentum. So it mechanically favours high-β semis/miners that have sold off, and fades low-vol "
"defensives/large-caps. The app's own CLAUDE.md flags this `low_vol` sign as the model's key fragility — it "
"could invert in a risk-off regime. **Where the factor model says BUY but posture says Bearish, the model is "
"betting on mean-reversion against the current downtrend — the lowest-confidence setups.**\n",
"## Master table (stocks by factor composite, ETFs last)\n",
"| # | Symbol | Factor | Posture | Agreement | Pctile | ExpFwd20d | Trend | RSI | MACD | ADX | 1mo% | β3mo | Driver |",
"|--:|---|:--:|:--:|:--:|--:|--:|:--:|--:|:--:|--:|--:|--:|:--:|"]
for i,(s,r) in enumerate(T.iterrows(),1):
    pc="—" if np.isnan(r.Pctile) else f"{r.Pctile:.0f}"
    ex="—" if r.Factor=="n/a" else f"{r.ExpFwd20d:+.1f}%"
    md.append(f"| {i} | **{s}**{r.Young} | {r.Factor} | {r.Posture} | {r.Agreement} | {pc} | {ex} | {r.Trend} | {r.RSI14:.0f} | {r.MACD[:4]} | {r.ADX14:.0f} | {r.Chg_1m:+.0f} | {r.Beta3mo:.1f} | {r.DomSector} |")
md.append("\n`*` limited history (young listing; 12-mo momentum factor neutralised).\n\n---\n\n## Detailed cards\n")
for i,(s,r) in enumerate(T.iterrows(),1):
    head=f"### {i}. {s} — {NAMES.get(s,s)}"
    if r.Factor=="n/a":
        head+=f" · Posture **{r.Posture}** (index ETF — factor model n/a)"
    else:
        head+=f" · Factor **{r.Factor}** / Posture **{r.Posture}** · _{r.Agreement}_"
    md.append(head+(f" · ⚠ limited history" if r.Young else "")+"\n")
    md.append(f"- **Price / return** — ${r.Price:,.2f} · 1w {r.Chg_1w:+.1f}% · 1mo {r.Chg_1m:+.1f}% · 3mo {r.Chg_3m:+.1f}% · at {r.Pos_52wk:.0f}% of the 52-wk range")
    md.append(f"- **Technical posture ({r.Posture})** — {tech_read(r)} ATR {r.ATR_pct:.1f}%/day.")
    if r.Factor!="n/a":
        md.append(f"- **Factor model ({r.Factor}, {r.Pctile:.0f}th pctile)** — composite {r.Composite:+.2f} → band {r.Band:.0f} (exp {r.ExpFwd20d:+.2f}% / 20d, {r.HitRate:.0f}% hit). Drivers: mom12-1 {sgn(r.z_mom12)} ({r.z_mom12:+.1f}z) · mom6-1 {sgn(r.z_mom6)} ({r.z_mom6:+.1f}z) · trend {sgn(r.z_trend)} ({r.z_trend:+.1f}z) · low-vol {sgn(r.z_lowvol)} ({r.z_lowvol:+.1f}z) · RS-sector {sgn(r.z_rss)} ({r.z_rss:+.1f}z) · turnover {sgn(r.z_turn)} ({r.z_turn:+.1f}z).")
        if r.Agreement.startswith("CONFLICT"):
            md.append(f"- **⚠ Conflict** — factor model ({r.Factor}) opposes the {r.Posture.lower()} price trend; this is a mean-reversion bet against momentum. Low confidence — wait for trend confirmation or treat as contrarian only.")
    md.append(f"- **Relationship / risk** — daily-3mo β(SPY) {r.Beta3mo:.1f} · intraday market R² {r.R2_intraday:.2f} · dominant driver **{r.DomSector}** · 3mo corr vs SPY {r.RS_SPY_3mo:+.2f} · VIX corr {r.VIXcorr:+.2f} ({'risk-on' if r.VIXcorr<-0.15 else 'defensive/inverse' if r.VIXcorr>0.10 else 'neutral'}).\n")
open(f"{OUT}/{DATE}-swing-per-symbol.md","w",encoding="utf-8").write("\n".join(md))
print("wrote", f"{OUT}/{DATE}-swing-per-symbol.md")

pd.set_option("display.width",260,"display.max_rows",60)
print("\n=== FACTOR SIGNAL COUNTS (stocks) ===")
print(T.loc[[s for s in T.index if s in STOCKS],"Factor"].value_counts().to_string())
print("\n=== POSTURE COUNTS (all) ==="); print(T.Posture.value_counts().to_string())
print("\n=== AGREEMENT ==="); print(T.Agreement.value_counts().to_string())
print("\n=== CONFIRMED setups (both lenses agree) ===")
conf=T[T.Agreement.str.startswith("Confirmed")]
print(conf[["Factor","Posture","Pctile","Trend","RSI14","MACD","Chg_1m","Chg_3m","DomSector"]].round(1).to_string())
print("\n=== CONFLICTS (low confidence) ===")
print(T[T.Agreement.str.startswith("CONFLICT")][["Factor","Posture","Trend","RSI14","MACD","Chg_1m","z_lowvol"]].round(1).to_string())
